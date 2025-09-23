from functools import reduce
from typing import Optional
from loguru import logger
import os
import json
import openpyxl


G  = 1e9
KB = 1<<10
MB = 1<<20
ico_list = ("in0", "in1", "out")
ico_m_list = ("IM", "CM", "OM")
zero_res = {
    "FLOPs": 0,
    "IOsize": 0,
    "m_act": 0,
    "m_weight": 0,
    "m_cache": 0,
    "m_other": 0,
    "t_TIU": 0,
    "t_DMA": 0,
    "t_MAX": 0,
}

def ceil(x:int|float) -> int:
    if isinstance(x, int):
        return x
    else:
        return int(x) + (1 if x - int(x) > 1e-3 else 0)
    
def align(x:int, n:int) -> int:
    """
        Align x to n_aling.
        Args:
            x (int): Value to align.
            n_aling (int): Align value.
        Returns:
            int: Aligned value.
    """
    return ceil(x/n) * n

def roundup_ratio(x:int|float, n:int) -> int:
    return x/align(x, n)

def rsum(*numbers):
    return reduce(lambda s,x: s + (1./x if x!=0 else 0), numbers, 0)

class TPUAttrs:
    device_name :str    = "sg2260" # Device name.
    DRAMBW      :float  = 546*0.9  # DRAM bandwidth.
    FLOP_power  :float  = 128      # FLOP power.
    NPU_NUM     :int    = 64       # NPU number.

    low_perf_DMA:float  = 0.5 # Low performance DMA.
    low_perf_TIU:float  = 0.8 # Low performance TIU.
    is_under_low:bool   = False # Under low performance.

    core_num    :int    = 8 # Core number.
    mem_local   :float  = 16 * 0.88 # Local memory size.
    v_CDMA      :float  = 32 # v_CDMA.
    v_ddr2l2    :float  = 68.25*0.9 # v_ddr2l2.
    v_ddr2local :float  = 68.25*0.9 # v_ddr2local.
    v_local2ddr :float  = 68.25*0.9 # v_local2ddr.
    v_local2l2  :float  = 128*0.9 # v_local2l2.
    v_l22local  :float  = 128*0.9 # v_l22local.
    v_SUM       :float  = 512 # v_SUM.
    lt_fire     :float  = 0.3 # lt_fire.
    lt_sync     :float  = 0.3 # lt_sync.

    mem_d1_div  :float  = 6  # Memory division for d1.

    @classmethod
    def update(cls, attrs:dict) -> None:
        for k, v in attrs.items():
            if hasattr(cls, k):
                setattr(cls, k, v)
        
        cls.ddr2l2    = cls.DRAMBW/cls.core_num
        cls.ddr2local = cls.DRAMBW/cls.core_num
        cls.local2ddr = cls.DRAMBW/cls.core_num
    
    @classmethod
    def setLow(cls, DMA_ratio=None, TIU_ratio=None) -> None:
        if not cls.is_under_low:
            cls.low_perf_DMA = cls.low_perf_DMA if DMA_ratio is None else DMA_ratio
            cls.low_perf_TIU = cls.low_perf_TIU if TIU_ratio is None else TIU_ratio
            cls.DRAMBW = cls.DRAMBW * cls.low_perf_DMA
            cls.FLOP_power = cls.FLOP_power * cls.low_perf_TIU
            cls.update({})
            cls.is_under_low = True
        else:
            if DMA_ratio is not None and DMA_ratio != cls.low_perf_DMA:
                cls.DRAMBW = cls.DRAMBW / cls.low_perf_DMA * DMA_ratio
                cls.low_perf_DMA = DMA_ratio
            if TIU_ratio is not None and TIU_ratio!= cls.low_perf_TIU:
                cls.FLOP_power = cls.FLOP_power / cls.low_perf_TIU * TIU_ratio
                cls.low_perf_TIU = TIU_ratio
            cls.update({})
    
    @classmethod
    def resetLow(cls) -> None:
        if cls.is_under_low:
            cls.DRAMBW = cls.DRAMBW / cls.low_perf_DMA
            cls.FLOP_power = cls.FLOP_power / cls.low_perf_TIU
            cls.update({})
            cls.is_under_low = False

class SuperParSelOpBase:
    """
        Basic class for super parameter selection operators.
        Args:
            matrix (dict|list): Shape info.
                in0 (list): Input shape.
                in1 (list): 2nd input or weight shape.
                out (list): Output shape.
            info: Other op info.
                matrix (dict|list): above.
                tp (int): TP.
                batch (int): Batch size.
                ds (int): Data size.
                ws (int): Weight size.
                IM (bool): Transfer in0.
                CM (bool): Transfer in1.
                OM (bool): Transfer out.
                NM (bool): Transfer nothing.
    """
    name  :str   = "op"  # Op name
    type  :str   = "base"# Op type
    tp    :int   = 1     # TP
    batch :int   = 1     # Batch size
    ds    :int   = 2     # Data size
    ws    :int   = 2     # Weight size
    IM    :bool  = True  # Transfer in0
    CM    :bool  = True  # Transfer in1
    OM    :bool  = True  # Transfer out
    NM    :bool  = False # Transfer nothing
    inBig :bool  = False # In big OP.
    mode  :str   = 'decode' # Mode: decode, prefill, train
    layer_num:int = 1     # Layer number
    moreAttr:bool = False # More attribute not pre-defined.

    run_cmd       :dict= {} # run cmd
    debug         :int = 0  # Debug level, 0: no debug, 1: print info, 2: print to excel.
    first_or_last :int = 0  # 1: first, -1: last, 0: nothing.
    last_res      :dict= {} # Last result.
    
    mac_utilization:float = 1.0 # MAC utilization.
    low_perf_DMA:float = 0.5 # Low performance DMA.
    low_perf_TIU:float = 0.8 # Low performance TIU.

    def __new__(cls, *args, **kwargs):
        for k, v in cls.__dict__.items():
            if isinstance(v, (dict, list)):
                v.clear()
        return super().__new__(cls)

    def __init__(self, info:dict) -> None:
        if not self.moreAttr and 'matrix' not in info:
            raise NotImplementedError(f"Matrix is not provided! {info=}")

        for k, v in info.items():
            if k=="matrix":
                if isinstance(v, dict):
                    for s in ico_list:
                        setattr(self, s, v[s] if s in v else [])
                elif isinstance(v, list):
                    self.in0 = self.in1 = self.out = v
                else:
                    raise NotImplementedError(f"Matrx type is not supported! Select from dict or list. {v=}")

            else:
                if self.moreAttr or hasattr(self, k):
                    if isinstance(v, int|float) and v < 0:
                        logger.warning(f"{k} is negative! {v=}")
                        continue
                    setattr(self, k, v)
                else:
                    raise NotImplementedError(f"Attribute {k} is not supported!")

        if "matrix" not in info:
            self.in0 = self.in1 = self.out = []

        if self.NM:
            self.IM = self.CM = self.OM = False
        elif self.inBig:
            # No I/O transfer in big op
            self.IM = self.OM = False

        self.checkInfo()

    def inferShape(self, shape:list):
        return [ self.getValue(x) for x in shape ]

    @property
    def op_name(self):
        return f"{self.__class__.__name__}.{self.name}"

    def __repr__(self):
        return f"{self.op_name}->{self.__class__.__name__}({self.inferShape(self.in0)}, {self.inferShape(self.in1)}, {self.inferShape(self.out)})"

    def checkInfo(self) -> None:
        pass

    def updateInfo(self, info:dict) -> None:
        for k, v in info.items():
            if hasattr(self, k):
                if isinstance(v, int|float) and v < 0:
                    logger.warning(f"{k} is negative! {v=}")
                    continue
                setattr(self, k, v)
            else:
                raise NotImplementedError(f"Attribute {k} is not defined!")

    def getValue(self, s:str) -> int:
        """
            Get value by string.
            Returns:
                int: Value.
        """
        num =  eval(s) if isinstance(s, str) else s
        return ceil(num)

    def getMatrixSize(self, shape:list) -> int:
        if shape:
            matrix_size = reduce(lambda x, y: x*y, (self.getValue(x) for x in shape))
            logger.debug(f"{self.op_name}.matrix_size {shape} = {self.inferShape(shape)} = {matrix_size}")
            return matrix_size
        else:
            return 0

    def getNparam(self) -> int:
        """
            Get trainable parameters.
        """
        return 0

    def forward(self) -> dict:
        return zero_res.copy()

    def backward(self) -> dict:
        return zero_res.copy()

    def __call__(self, mode:str='decode', tp:int=1, batch:int=1, run_cmd:dict={}) -> dict:
        """
            Forward function.
            Returns: [G] [MB] [us]
                dict: {
                    FLOPs,
                    IOsize,
                    m_act,
                    m_weight,
                    m_cache,
                    m_other,
                    t_TIU,
                    t_DMA,
                    t_MAX,
                }
        """
        self.setModeTPbatch(mode, tp, batch, run_cmd)
        if mode in ('decode', 'prefill'):
            self.last_res = self.forward()
        elif mode == 'backward':
            self.last_res = self.backward()
        elif mode == 'train':
            self.last_res = self.forward()
            b_res = self.backward()
            for k in b_res:
                self.last_res[k] += b_res[k]
        else:
            raise NotImplementedError(f"Mode {mode} is not supported!")

        if hasattr(self, "sub_ops"):
            for op in getattr(self, "sub_ops"):
                sub_res = op(mode, tp, batch, run_cmd)
                for k in zero_res:
                    self.last_res[k] += getattr(op, "layer_num", 1) * sub_res[k]
                    logger.debug(f"{self.op_name}.{k} += {getattr(op, 'layer_num', 1)} * {sub_res[k]} = {self.last_res[k]}")

        if self.last_res['t_TIU'] <= 0 and self.last_res['FLOPs'] > 0:
            self.last_res['t_TIU'] = self.last_res['FLOPs'] / self.mac_utilization / TPUAttrs.FLOP_power / 1024 * 1e6
            logger.debug(f"{self.op_name}.t_TIU = {self.last_res['FLOPs']} / {self.mac_utilization} / {TPUAttrs.FLOP_power} / 1024 * 1e6 = {self.last_res['t_TIU']}")

        if self.last_res['t_DMA'] <= 0 and self.last_res['IOsize'] > 0:
            self.last_res['t_DMA'] = self.last_res['IOsize'] / KB / TPUAttrs.DRAMBW * 1e6
            logger.debug(f"{self.op_name}.t_DMA = {self.last_res['IOsize']} / KB / {TPUAttrs.DRAMBW} * 1e6 = {self.last_res['t_DMA']}")

        if self.type != "LLM" and not getattr(self, "OptimizedTime", False):
            self.last_res['t_MAX'] = max(self.last_res['t_TIU'], self.last_res['t_DMA'])
            logger.debug(f"{self.op_name}.t_MAX = max({self.last_res['t_TIU']}, {self.last_res['t_DMA']}) = {self.last_res['t_MAX']}")

        self.run_cmd = {}
        return self.last_res

    def setModeTPbatch(self, mode:str, tp:int, batch:int, run_cmd:dict={}) -> None:
        """
            Set mode, tp and batch size. And run args.
        """
        self.mode  = mode
        self.tp    = tp
        self.batch = batch
        self.run_cmd = run_cmd
        
        # 0: low, 1: high
        perf_mode = run_cmd.get('perf_mode', 1)
        if perf_mode == 0:
            TPUAttrs.setLow(self.low_perf_DMA, self.low_perf_TIU)
        else:
            TPUAttrs.resetLow()

    def to_dict(self):
        res = {
            "in0": self.in0,
            "in1": self.in1,
            "out": self.out,
            "in_big": self.inBig,
            "inference": self.last_res,
            "layer_num":  getattr(self, "layer_num", 1)
        }
        
        if hasattr(self, "sub_ops"):
            res['sub_ops'] = []
            for op in getattr(self, "sub_ops"):
                res["sub_ops"].append(op.to_dict())
        
        return res

    def dump2excel(self, workbook:openpyxl.Workbook=None, outname:str="") -> None:
        """
            Print op info to excel.
        """
        if isinstance(workbook, openpyxl.Workbook):
            ws = workbook.active
            in0 = [ self.getValue(x) for x in self.in0 ] if self.in0 else ["",]*4
            in1 = [ self.getValue(x) for x in self.in1 ] if self.in1 else ["",]*4
            out = [ self.getValue(x) for x in self.out ] if self.out else ["",]*4
            ws.append(
                ["| " if self.inBig else "-" if self.type == "Big" else "", self.name, self.type,] + 
                in0 + in1 + out +
                [self.last_res['FLOPs'], self.last_res['t_TIU'], self.last_res['IOsize'], self.last_res['t_DMA'], self.last_res['t_MAX'],] + 
                [self.last_res['m_act'], self.last_res['m_weight'], self.last_res['m_cache'], self.last_res['m_other']]
            )

            if outname:
                center_align = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = center_align
                workbook.save(outname)
                logger.info(f"Excel file saved to {outname}")

        elif workbook is None:
            logger.info(f"Generating excel file...")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"SPS_{self.mode}"
            titles = ["", "", "", "in0", "", "", "", "in1", "", "", "", "out", "", "", "", "理论性能", "", "", "", "", "内存分析", "", "", ""]
            ws.append(titles)
            merge_col_L = 1
            for i in range(merge_col_L+1, len(titles)+1):
                if i == len(titles) or titles[i]:
                    merge_col_R = i
                    ws.merge_cells(start_row=1, start_column=merge_col_L, end_row=1, end_column=merge_col_R)
                    merge_col_L = i+1

            ws.append(["BigOpName", "OpName", "OpType", "N", "C", "H", "W", "N", "C", "H", "W", "N", "C", "H", "W", "FLOPs [G]", "TIU [us]", "IO size [MB]", "DMA [us]", "Total [us]", "Activation [MB]", "Weight [MB]", "Cache [MB]", "Other [MB]"])
            self.dump2excel(wb, outname)

        else:
            raise NotImplementedError(f"Workbook type is not supported! {workbook=}")

class SPS_MM2(SuperParSelOpBase):
    """
        Matrix multiplication for super parameter selection.
            noCalTime (bool): No calculation time.
            isParam (bool): Has trainable parameters.
            s_noMem (str): No memory for in0/in1/out.
            w_multi (int): Weight multiplier.
    """
    noCalTime :bool = False # No calculation time.
    isParam   :bool = True  # Trainable parameter.
    s_noMem   :str  = ''    # No memory for in0/in1/out.
    w_multi   :int  = 1     # Weight multiplier.
    transpose :bool = False # MM2_NT.
    OptimizedTime :bool = True # Optimized t_MAX.
    quant_method :str|list[str] = None  # ["w4a16", "fp8"]
    def __init__(self, info:dict) -> None:
        super().__init__(info)
        if not self.isParam and self.inBig:
            self.CM = False

        if self.inBig:
            self.OptimizedTime = False

    def checkInfo(self):
        """
            Rough check.
        """
        # TODO: support shape alignment.
        if not self.transpose:
            if self.in0[-1] != self.in1[-2]:
                raise NotImplementedError(f"Expect in0_W = in1_H, but {self.in0[-1]} != {self.in1[-2]}")
            elif self.in1[-1] != self.out[-1]:
                raise NotImplementedError(f"Expect in1_W = out_W, but {self.in1[-1]} != {self.out[-1]}")
        else: 
            pass
    
    def getNparam(self) -> int:
        """
            Get trainable parameters.
        """
        if self.isParam:
            return self.getMatrixSize(self.in1) * self.w_multi
        else:
            return 0

    def forward(self):
        # X_l+1 = X_l @ W_l^T  ~~>  out = in0 @ in1  ~~>  [a,c] ~ [a,b] @ [b,c]  ~~> FLOPs: 2abc  ~~> IOsize: ab+bc+ac  !!!! be careful if MM is in a fusion operator 
        res = zero_res.copy()

        if not self.noCalTime:
            # FLOPs = 2*M*N*K
            mid_shape = self.in1[-2] if not self.transpose else self.in1[-1]
            res['FLOPs'] = 2 * self.getMatrixSize(self.out)*self.getValue(mid_shape) / G
            logger.debug(f"{self.op_name}.FLOPs = 2 * {self.getMatrixSize(self.out)} * {self.getValue(mid_shape)} / G = {res['FLOPs']}")

            # Update MAC utilization
            if TPUAttrs.NPU_NUM > 0:
                ### TODO: Update this, TPU v7.1指令集 13.1:TPU性能公式
                tmpmaxS = -1
                for i in (1,2,0): # C, H, N
                    if self.getValue(self.out[i]) > tmpmaxS:
                        tmpmaxS = self.getValue(self.out[i])
                self.mac_utilization = tmpmaxS/TPUAttrs.NPU_NUM if tmpmaxS<TPUAttrs.NPU_NUM else 1
                if self.mac_utilization < 0.9:
                    logger.debug(f"{self.op_name}.mac_utilization = {tmpmaxS}/{TPUAttrs.NPU_NUM} = {self.mac_utilization}")
        
            res['t_TIU'] = res['FLOPs'] / self.mac_utilization / TPUAttrs.FLOP_power / 1024 * 1e6
            logger.debug(f"{self.op_name}.t_TIU = {res['FLOPs']} / {self.mac_utilization} / {TPUAttrs.FLOP_power} / 1024 * 1e6 = {res['t_TIU']}")

        if not self.NM:
            if self.isParam:
                # save activation for trainable parameters
                if self.mode in ('decode', 'prefill') and self.inBig:
                    res['m_act'] = 0
                else:
                    matrix_size = self.getMatrixSize(self.in0)
                    res['m_act']    = self.ds * matrix_size / MB #* CoreNum
                    # TODO: check if output / CoreNum is needed
                    logger.debug(f"{self.op_name}.m_act = {self.ds} * {matrix_size} / MB  = {res['m_act']}")

                matrix_size = self.getMatrixSize(self.in1)
                #  qunat method
                quant = self.run_cmd.get("quant_method", 'f16')
                logger.debug(f"{self.op_name}.quant_method = {quant}, {self.run_cmd=}")
                if quant!='f16' and self.quant_method:
                    if quant == 'w4a16' and 'w4a16' in self.quant_method:
                        # group_size = 128 #  w int4 + scale f16/gs + zp int4/gs
                        scale = 1/2 + 2/128 + 1/2/128
                    elif quant == 'fp8' and 'fp8' in self.quant_method:
                        if self.ws == 1:
                            logger.warning("FP8 weight is set when ws=1, which means fp8 weight is used.")
                        scale = 1
                    else:
                        raise NotImplementedError(f"Quant method {quant} is not supported!")
                    res['m_weight'] = scale   * matrix_size / MB
                    logger.debug(f"{self.op_name}.m_weight = {scale  } * {matrix_size} / MB = {res['m_weight']}")
                else:
                    res['m_weight'] = self.ws * matrix_size / MB
                    logger.debug(f"{self.op_name}.m_weight = {self.ws} * {matrix_size} / MB = {res['m_weight']}")

            # IOsize = in0 + in1 + out
            useIO = [ getattr(self, s) for s in ico_m_list ]
            data_sizes = [self.ds, self.ws, self.ds]
            io_ds = [0,0,0]
            for i,s in enumerate(ico_list):
                if useIO[i]:
                    if i==1 and self.isParam:
                        io_ds[i] = res['m_weight']
                    else:
                        io_ds[i] = data_sizes[i] * self.getMatrixSize(getattr(self, s)) / MB
                    res['IOsize'] += io_ds[i]

            res['t_DMA'] = (res['IOsize'] / KB / TPUAttrs.DRAMBW) * 1e6
            logger.debug(f"{self.op_name}.t_DMA = {res['IOsize']} / KB / {TPUAttrs.DRAMBW} * 1e6 = {res['t_DMA']}")

            # Optimized t_MAX ## in local memory, fp8/int4 -> fp16
            if self.OptimizedTime:
                tmp_total_mem = io_ds[0] + io_ds[1]/TPUAttrs.core_num + io_ds[2]/TPUAttrs.core_num
                if tmp_total_mem < TPUAttrs.mem_local:
                    res['t_MAX'] = (io_ds[0]/TPUAttrs.core_num/TPUAttrs.v_ddr2l2 + io_ds[0]/TPUAttrs.v_l22local + io_ds[1]/TPUAttrs.core_num/TPUAttrs.v_ddr2local + io_ds[2]/TPUAttrs.core_num/TPUAttrs.v_local2ddr)*1e6/KB + res['t_TIU']
                    logger.debug(f"{self.op_name}.t_MAX = ({io_ds[0]}/{TPUAttrs.core_num}/{TPUAttrs.v_ddr2l2} + {io_ds[0]}/{TPUAttrs.v_l22local} + {io_ds[1]}/{TPUAttrs.core_num}/{TPUAttrs.v_ddr2local} + {io_ds[2]}/{TPUAttrs.core_num}/{TPUAttrs.v_local2ddr})*1e6/{KB} + {res['t_TIU']} = {res['t_MAX']}")
                else:
                    d1_core = io_ds[1] / TPUAttrs.core_num
                    res['t_MAX'] = (io_ds[1]/TPUAttrs.core_num/TPUAttrs.v_ddr2local)*1e6/KB + res['t_TIU'] * ( (align(d1_core, TPUAttrs.mem_d1_div)-d1_core)/d1_core if d1_core>0 and res['t_TIU'] < res['t_DMA'] else 1)
                    logger.debug(f"{self.op_name}.t_MAX = ({io_ds[1]}/{TPUAttrs.core_num}/{TPUAttrs.v_ddr2local})*1e6/{KB} + {res['t_TIU']} * ({(align(d1_core, TPUAttrs.mem_d1_div)-d1_core)/d1_core if d1_core>0 and res['t_TIU'] < res['t_DMA'] else 1}) = {res['t_MAX']}")

        # MoE weight multiplier
        if self.w_multi > 1:
            res['m_weight'] *= self.w_multi
            logger.debug(f"{self.op_name}.m_weight *= {self.w_multi} = {res['m_weight']}")

        return res

    def backward(self):
        # GradX_l-1 = GradX_l @ W       ~~>  in0 ~ out @ in1^T  ~~>  [a,b] ~ [a,c] @ [c,b]  ~~> FLOPs: 2abc  ~~> IOsize: ab+bc+ac
        # GradW_l^T = X_l-1^T @ GradX_l ~~>  in1 ~ in0^T @ out  ~~>  [b,c] ~ [b,a] @ [a,c]  ~~> FLOPs: 2abc  ~~> IOsize: ab+bc+ac
        res = zero_res.copy()

        if not self.noCalTime:
            # FLOPs = 2*2*M*N*K
            mid_shape = self.in1[-2] if not self.transpose else self.in1[-1]
            res['FLOPs'] = 2 * 2 * self.getMatrixSize(self.out)*self.getValue(mid_shape) / G

            # NPU utilization, train seq usually > NPU_NUM
            pass

        # Each parameter has 16 bytes for gradient
        res['m_other'] = 16 * self.getNparam() / MB
        logger.debug(f"{self.op_name}.m_other = 16 * {self.getNparam()} / MB = {res['m_other']}")
        
        if not self.NM:
            if self.isParam:
                # save activation for trainable parameters
                res['m_act'] = self.ds * (self.getMatrixSize(self.in0) + self.getMatrixSize(self.out)) / MB
                logger.debug(f"{self.op_name}.m_act = {self.ds} * ({self.getMatrixSize(self.in0)} + {self.getMatrixSize(self.out)}) / MB = {res['m_act']}")

            # Weight has been considered in forward
        
            # IOsize = in0 + in1 + out
            useIO = [ getattr(self, s) for s in ico_m_list ]
            data_sizes = [self.ds, self.ds, self.ds] ### Mix precision training, use f16 as default
            for i,s in enumerate(ico_list):
                if useIO[i]:
                    res['IOsize'] += data_sizes[i] * self.getMatrixSize(getattr(self, s)) / MB
        return res

class SPS_MM2_NT(SPS_MM2):
    transpose:bool = True

class SPS_AR(SuperParSelOpBase):
    """
        TODO: AI generated, need to be verified.
        Add, Mul, Div operations for super parameter selection.
        Args:
            info (dict): Operation info.
                matrix (dict|list): Shape info.
                flops_per_element (int): FLOPs per element, default 2.
    """
    flops_per_element: int = 1  # Default FLOPs per element for AR operations
    
    def forward(self):
        res = zero_res.copy()
        
        # Step 1: Calculate FLOPs based on output size and flops per element
        res['FLOPs'] = self.flops_per_element * self.getMatrixSize(self.out) / G
        logger.debug(f"{self.op_name}.FLOPs = {self.flops_per_element} * {self.getMatrixSize(self.out)} / G = {res['FLOPs']}")
        
        # Step 2: Calculate TIU time based on FLOPs
        res['t_TIU'] = res['FLOPs'] / self.mac_utilization / TPUAttrs.FLOP_power / 1024 * 1e6
        logger.debug(f"{self.op_name}.t_TIU = {res['FLOPs']} / {self.mac_utilization} / {TPUAttrs.FLOP_power} / 1024 * 1e6 = {res['t_TIU']}")
        
        # Step 3: Calculate memory usage and IO size
        # Memory for input and output
        res['m_act'] = self.ds * self.getMatrixSize(self.in0) / MB
        logger.debug(f"{self.op_name}.m_act = {self.ds} * {self.getMatrixSize(self.in0)} / MB = {res['m_act']}")
        
        # IO size includes input and output
        useIO = [getattr(self, s) for s in ico_m_list]
        data_sizes = [self.ds, self.ds, self.ds]
        for i, s in enumerate(ico_list):
            if useIO[i]:
                res['IOsize'] += data_sizes[i] * self.getMatrixSize(getattr(self, s)) / MB
                logger.debug(f"{self.op_name}.IOsize += {data_sizes[i]} * {self.getMatrixSize(getattr(self, s))} / MB = {res['IOsize']}")
        
        # Step 4: Calculate DMA time based on IO size
        res['t_DMA'] = res['IOsize'] / KB / TPUAttrs.DRAMBW * 1e6
        logger.debug(f"{self.op_name}.t_DMA = {res['IOsize']} / KB / {TPUAttrs.DRAMBW} * 1e6 = {res['t_DMA']}")
        
        # Step 5: Total time is max of TIU and DMA
        res['t_MAX'] = max(res['t_TIU'], res['t_DMA'])
        logger.debug(f"{self.op_name}.t_MAX = max({res['t_TIU']}, {res['t_DMA']}) = {res['t_MAX']}")
        
        return res
    
    def backward(self):
        res = zero_res.copy()
        
        # Step 1: Calculate FLOPs for backward pass (usually 2x forward)
        res['FLOPs'] = 2 * self.flops_per_element * self.getMatrixSize(self.out) / G
        logger.debug(f"{self.op_name}.FLOPs = 2 * {self.flops_per_element} * {self.getMatrixSize(self.out)} / G = {res['FLOPs']}")
        
        # Step 2: Calculate TIU time for backward
        res['t_TIU'] = res['FLOPs'] / self.mac_utilization / TPUAttrs.FLOP_power / 1024 * 1e6
        logger.debug(f"{self.op_name}.t_TIU = {res['FLOPs']} / {self.mac_utilization} / {TPUAttrs.FLOP_power} / 1024 * 1e6 = {res['t_TIU']}")
        
        # Step 3: Calculate memory usage and IO size for backward
        if not self.NM:
            # Memory for input and output gradients
            res['m_act'] = self.ds * (self.getMatrixSize(self.in0) + self.getMatrixSize(self.out)) / MB
            logger.debug(f"{self.op_name}.m_act = {self.ds} * ({self.getMatrixSize(self.in0)} + {self.getMatrixSize(self.out)}) / MB = {res['m_act']}")
            
            # IO size includes input and output gradients
            useIO = [getattr(self, s) for s in ico_m_list]
            data_sizes = [self.ds, self.ds, self.ds]
            for i, s in enumerate(ico_list):
                if useIO[i]:
                    res['IOsize'] += data_sizes[i] * self.getMatrixSize(getattr(self, s)) / MB
                    logger.debug(f"{self.op_name}.IOsize += {data_sizes[i]} * {self.getMatrixSize(getattr(self, s))} / MB = {res['IOsize']}")
            
            # Step 4: Calculate DMA time for backward
            res['t_DMA'] = res['IOsize'] / KB / TPUAttrs.DRAMBW * 1e6
            logger.debug(f"{self.op_name}.t_DMA = {res['IOsize']} / KB / {TPUAttrs.DRAMBW} * 1e6 = {res['t_DMA']}")
            
            # Step 5: Total time is max of TIU and DMA
            res['t_MAX'] = max(res['t_TIU'], res['t_DMA'])
            logger.debug(f"{self.op_name}.t_MAX = max({res['t_TIU']}, {res['t_DMA']}) = {res['t_MAX']}")
        
        return res

class SPS_Mix(SuperParSelOpBase):
    """
        Mix for super parameter selection. RMSNorm
        kwargs: Other op info:
            isParam (bool): Has trainable parameters.
    """
    noCalTime :bool = False # No calculation time.
    isParam   :bool = True  # Trainable parameter.
    def getNparam(self) -> int:
        """
            Get trainable parameters.
        """
        if self.isParam:
            return self.getMatrixSize(self.in1)
        else:
            return 0
        
    def forward(self):
        res = zero_res.copy()
        # Mix op usually t_TIU << t_DMA
        if not self.noCalTime:
            # FLOPs ~ 4*data_size
            res['FLOPs'] = 4 * self.getMatrixSize(self.out) / G

        if not self.NM:
            if self.isParam:
                # save activation for trainable parameters
                ### TODO: Check if RMSNorm needs in0 as activation
                # if self.mode in ('decode', 'prefill') and self.inBig:
                res['m_act'] = 0
                # else:
                #     res['m_act'] = self.ds * self.getMatrixSize(self.in0) / MB 
                matrix_size = self.getMatrixSize(self.in1)
                res['m_weight']  = self.ws * matrix_size / MB
                logger.debug(f"{self.op_name}.m_weight = {self.ws} * {matrix_size} / MB = {res['m_weight']}")

        # IOsize = in0 + in1 + out
        useIO = [ getattr(self, s) for s in ico_m_list ]
        data_sizes = [self.ds, self.ws, self.ds]
        for i,s in enumerate(ico_list):
            if useIO[i]:
                if i==1 and self.isParam:
                    res['IOsize'] += res['m_weight']
                else:
                    res['IOsize'] += data_sizes[i] * self.getMatrixSize(getattr(self, s)) / MB
        
        return res
    
    def backward(self):
        res = zero_res.copy()
        # Mix op usually t_TIU << t_DMA
        if not self.noCalTime:
            # TODO: update scale or keep as t_TIU << t_DMA
            res['FLOPs'] = 8 * self.getMatrixSize(self.out) / G

        # Each parameter has 16 bytes for gradient
        res['m_other'] = 16 * self.getNparam() / MB
        logger.debug(f"{self.op_name}.m_other = 16 * {self.getNparam()} / MB = {res['m_other']}")

        if not self.NM:
            if self.isParam:
                # save activation for trainable parameters
                res['m_act'] = self.ds * (self.getMatrixSize(self.in0) + self.getMatrixSize(self.out)) / MB
                logger.debug(f"{self.op_name}.m_act = {self.ds} * ({self.getMatrixSize(self.in0)} + {self.getMatrixSize(self.out)}) / MB = {res['m_act']}")
        
        # IOsize = in0 + in1 + out
        useIO = [ getattr(self, s) for s in ico_m_list ]
        data_sizes = [self.ds, self.ws, self.ds]
        for i,s in enumerate(ico_list):
            if useIO[i]:
                res['IOsize'] += data_sizes[i] * self.getMatrixSize(getattr(self, s)) / MB
        
        return res
    
class SPS_ConvND(SuperParSelOpBase):
    """
        ConvXD for super parameter selection.
    """
    group:    int  = 1     # Group number.
    stride:   list = None  # Stride.
    padding:  list = None  # Padding.
    kernel:   list = None  # Kernel size.
    dilation: int  = 1     # Dilation.
    bias:     bool = False # Bias.

    opd0_arange_cycle: int = 1 # OpD0 arange cycle.
    def checkInfo(self):
        """
            in0: [batch, iC,       **Length]
            in1: [   oC, iC/group, **kernel]
            Check out size.
        """
        N = len(self.in0) - 2
        if not (len(self.in0) == len(self.in1) == len(self.out) == N + 2):
            raise NotImplementedError(f"Expect in0, in1, out have same length, but {len(self.in0)=}, {len(self.in1)=}, {len(self.out)=}, {N=}")
        self.dimention = N

        if self.padding is None:
            self.padding = [0,]*(2*N)
        elif len(self.padding) == N:
            self.padding = [x for x in self.padding for _ in range(2)]
        else:
            raise NotImplementedError(f"Padding length = {len(self.padding)} is not supported for Conv{N}D! {self.padding=}")
        
        if self.kernel is None:
            self.kernel = self.in1[2:]
        
        if self.stride is None:
            self.stride = [1,]*N
        
        if self.in0[1] % self.group != 0:
            raise NotImplementedError(f"Expect in0_C % group == 0, but {self.in0[1]} % {self.group} != 0")
    
        ### TODO: Update opd0_arange_cycle according to matrix size. # TPU1686/sg2260/cmodel/src/get_atomic_profile.cpp

    def getNparam(self):
        return self.getMatrixSize(self.in1)

    def forward(self):
        res = zero_res.copy()
        # For each element in output, it has about (2 * kernel_size * iC/group) FLOPs
        res['FLOPs'] = 2 * self.getMatrixSize(self.out) * self.getMatrixSize(self.kernel) * self.in0[1] / self.group / G
        logger.debug(f"{self.op_name}.FLOPs = 2 * {self.getMatrixSize(self.out)} * {self.getMatrixSize(self.kernel)} * {self.in0[1]} / {self.group} / G = {res['FLOPs']}")
        
        # Update MAC utilization
        self.mac_utilization = roundup_ratio(self.in0[1], 32) * roundup_ratio(self.out[1], TPUAttrs.NPU_NUM) / self.opd0_arange_cycle
        if self.mac_utilization < 0.9 or self.mac_utilization > 1.0:
            logger.debug(f"{self.op_name}.mac_utilization = {roundup_ratio(self.in0[1], 32)} * {roundup_ratio(self.out[1], TPUAttrs.NPU_NUM)} / {self.opd0_arange_cycle} = {self.mac_utilization}")
        
        res['m_act'] = self.ds * self.getMatrixSize(self.in0) / MB
        logger.debug(f"{self.op_name}.m_act = {self.ds} * {self.getMatrixSize(self.in0)} / MB = {res['m_act']}")

        # Consider 32ic
        res['m_weight'] = self.ws * self.getMatrixSize(self.in1) * align(self.in1[1], 32)/self.in1[1] / MB
        logger.debug(f"{self.op_name}.m_weight = {self.ws} * {self.getMatrixSize(self.in1)} / MB = {res['m_weight']}")

        # IOsize = in0 + in1 + out
        useIO = [ getattr(self, s) for s in ico_m_list ]
        data_sizes = [self.ds, self.ws, self.ds]
        for i,s in enumerate(ico_list):
            if useIO[i]:
                if i==1:
                    res['IOsize'] += res['m_weight']
                else:
                    res['IOsize'] += data_sizes[i] * self.getMatrixSize(getattr(self, s)) / MB

        return res

class superParSelBigOpBase(SuperParSelOpBase):
    """
        Big op basic class for super parameter selection.
        Args: IO & sub op info:
            matrix (dict|list): Shape info.
                in0 (list): Input shape.
                in1 (list): 2nd input or weight shape.
                out (list): Output shape.
            info:
                matrix (dict|list): above.
                cache (dict): Cache info.
                sub_ops (list): Sub op info.
    """
    cache   : dict = {}
    sub_info: list = []
    sub_ops : list[SuperParSelOpBase] = []
    def __init__(self, info:dict):
        super().__init__(info)

        self.sub_ops: list[SuperParSelOpBase] = []
        for op in self.sub_info:
            op['inBig'] = True
            op['low_perf_DMA'] = self.low_perf_DMA
            op['low_perf_TIU'] = self.low_perf_TIU
            self.sub_ops.append(gen_SPS_OP(op))

    def getNparam(self):
        return reduce(lambda x, y: x+y, (getattr(op, "layer_num", 1) * op.getNparam() for op in self.sub_ops))

    def forward(self):
        res = zero_res.copy()

        # KV_cache etc.
        for s in self.cache:
            res['m_cache'] += self.ds * self.getMatrixSize(self.cache[s]) / MB
            logger.debug(f"{self.op_name}.m_cache += {self.ds} * {self.getMatrixSize(self.cache[s])} / MB = {res['m_cache']}")

        # IOsize = in0 + in1 + out + cache
        data_sizes = [self.ds, self.ws, self.ds]
        for i,s in enumerate(ico_list):
            res['IOsize'] += data_sizes[i] * self.getMatrixSize(getattr(self, s)) / MB
            logger.debug(f"{self.op_name}.IOsize += {data_sizes[i]} * {self.getMatrixSize(getattr(self, s))} / MB = {res['IOsize']}")
        res['IOsize'] += res['m_cache']
        logger.debug(f"{self.op_name}.IOsize += {res['m_cache']} = {res['IOsize']}")

        res['t_DMA']  += res['IOsize'] / KB / TPUAttrs.DRAMBW * 1e6
        res['IOsize_bk'] = res['IOsize']
        res['t_DMA_bk'] = res['t_DMA'] 
        return res
    
    def backward(self):
        res = zero_res.copy()

        # IOsize = in0 + grad(in0) + in1 + grad(out)
        data_sizes = [2*self.ds, self.ws, self.ds]
        for i,s in enumerate(ico_list):
            res['IOsize'] += data_sizes[i] * self.getMatrixSize(getattr(self, s)) / MB
            logger.debug(f"{self.op_name}.IOsize += {data_sizes[i]} * {self.getMatrixSize(getattr(self, s))} / MB = {res['IOsize']}")

        res['t_DMA']  += res['IOsize'] / KB / TPUAttrs.DRAMBW * 1e6
        res['IOsize_bk'] = res['IOsize']
        res['t_DMA_bk'] = res['t_DMA'] 
        return res
        
    def dump2excel(self, workbook:openpyxl.Workbook=None, outname:str="") -> None:
        if isinstance(workbook, openpyxl.Workbook):
            ws = workbook.active
            in0 = [ self.getValue(x) for x in self.in0 ] if self.in0 else ["",]*4
            in1 = [ self.getValue(x) for x in self.in1 ] if self.in1 else ["",]*4
            out = [ self.getValue(x) for x in self.out ] if self.out else ["",]*4
            ws.append(
                ["|+" if self.inBig else self.name, self.name, self.type,] +
                in0 + in1 + out +
                ["-", "-", self.last_res['IOsize_bk'], self.last_res['t_DMA_bk']] + ["-"]*5 
            )

            for op in self.sub_ops:
                op.dump2excel(workbook)

        super().dump2excel(workbook, outname)

class SPS_AllReduce(SuperParSelOpBase):
    def forward(self):
        res = zero_res.copy()
        if self.tp == 1:
            return res
        
        matrix_size = self.getMatrixSize(self.in0)
        data_GBus   = self.ds * matrix_size * 1e-3
        data_GBus_per_tp = data_GBus / self.tp
        ppl_size = MB/2
        n_ppl = ceil(matrix_size/self.tp/ppl_size)
        if n_ppl > 1:
            logger.debug(f"{self.name}.n_ppl = ceil({matrix_size}/{self.tp}/{ppl_size}) = {n_ppl}")

        if self.tp == 2:
            if n_ppl <= 1:
                res['t_DMA'] = data_GBus_per_tp * rsum(TPUAttrs.v_local2l2, TPUAttrs.v_CDMA, TPUAttrs.v_SUM, TPUAttrs.v_l22local) + TPUAttrs.lt_fire + TPUAttrs.lt_sync
            else:
                ppl_data_GBus = self.ds * ppl_size * 1e-3
                t0 = ppl_data_GBus * rsum(TPUAttrs.v_l22local)
                t1 = ppl_data_GBus * rsum(TPUAttrs.v_CDMA, TPUAttrs.v_SUM) + TPUAttrs.lt_fire + TPUAttrs.lt_sync
                t2 = ppl_data_GBus * rsum(TPUAttrs.v_l22local)
                res['t_DMA'] = t0 + t1*n_ppl + t2
                logger.debug(f"{self.name}.t_DMA = {t0}+{t1}*{n_ppl}+{t2} = {res['t_DMA']}")
        elif self.tp % 2 == 0:
            if n_ppl <= 1:
                t0 =                 data_GBus_per_tp * rsum(TPUAttrs.v_local2l2)
                t1 = (self.tp/2-1) *(data_GBus_per_tp * rsum(TPUAttrs.v_CDMA,   TPUAttrs.v_SUM) +   TPUAttrs.lt_fire +   TPUAttrs.lt_sync)
                t2 =                 data_GBus_per_tp * rsum(TPUAttrs.v_CDMA, 2*TPUAttrs.v_SUM) + 2*TPUAttrs.lt_fire + 2*TPUAttrs.lt_sync
                t3 = (self.tp/2-1) *(data_GBus_per_tp * rsum(TPUAttrs.v_CDMA)                   +   TPUAttrs.lt_fire +   TPUAttrs.lt_sync)
                t4 =                 data_GBus_per_tp * rsum(TPUAttrs.v_l22local)
                res["t_DMA"] = t0 + t1 + t2 + t3 + t4
                logger.debug(f"{self.name}.t_DMA = {t0}+{t1}+{t2}+{t3}+{t4} = {res['t_DMA']}")
            else:
                ppl_data_GBus = self.ds * ppl_size * 1e-3
                t0 =                 ppl_data_GBus * rsum(TPUAttrs.v_local2l2)
                t1 = (self.tp/2-1) *(ppl_data_GBus * rsum(TPUAttrs.v_CDMA,   TPUAttrs.v_SUM) +   TPUAttrs.lt_fire +   TPUAttrs.lt_sync)
                t2 =                 ppl_data_GBus * rsum(TPUAttrs.v_CDMA, 2*TPUAttrs.v_SUM) + 2*TPUAttrs.lt_fire + 2*TPUAttrs.lt_sync
                t3 = (self.tp/2-1) *(ppl_data_GBus * rsum(TPUAttrs.v_CDMA)                   +   TPUAttrs.lt_fire +   TPUAttrs.lt_sync)
                t4 =                 ppl_data_GBus * rsum(TPUAttrs.v_l22local)
                res["t_DMA"] = t0 + (t1 + t2 + t3) * n_ppl + t4
                logger.debug(f"{self.name}.t_DMA = {t0}+({t1}+{t2}+{t3})*{n_ppl}+{t4} = {res['t_DMA']}")
        else:
            raise NotImplementedError(f"TPU number {self.tp} is odd!")
        
        return res

SPS_OP_table = {
    "MM2":    SPS_MM2,
    "MM2_NT": SPS_MM2_NT,
    "Mix":    SPS_Mix,
    "AR":     SPS_AR,
    "ConvND": SPS_ConvND,

    "AllReduce":   SPS_AllReduce,
    
    "Big":    superParSelBigOpBase
}

def gen_SPS_OP(alg:dict):
    """
        Generate SPS operator.
        Args:
            alg (dict): Algorithm info.
        Returns:
            dict: SPS operator.
    """
    if 'type' not in alg:
        raise NotImplementedError(f"No type info! {alg=}")
    
    if alg['type'] not in SPS_OP_table:
        raise NotImplementedError(f"Type {alg['type']} is not supported!")
    
    return SPS_OP_table[alg['type']](alg)

def get_SPS_LLM_op(raw_layers:list) -> list:
    if any(  isinstance(layer, str) or (isinstance(layer, dict) and hasattr(layer, "op")) for layer in raw_layers ):
        with open("examples/SPS_LLM_ops.json", 'r') as f2:
            pre_defined_ops = json.load(f2)
        
        def set_sub_attrs(d:dict, kbs:str, v):
            if ';' not in kbs:
                d[kbs] = v

            else:
                kb, ex_kb = kbs.split(';', 1)
                if ':' in kb:
                    k, bSet = kb.split(':')
                    bSet = [ False if x=='0' else True for x in bSet ]
                else:
                    k = kb
                    bSet = [ True for _ in range(len(d[k])) ]

                for i,b in enumerate(bSet):
                    if b:
                        set_sub_attrs(d[k][i], ex_kb, v)

        from copy import deepcopy
        inst_layers = []
        for layer in raw_layers:
            if isinstance(layer, str):
                inst_layers.append(deepcopy(pre_defined_ops[layer]))

            elif isinstance(layer, dict):
                if "op" not in layer:
                    inst_layers.append(layer)
                else:
                    origin_op = deepcopy(pre_defined_ops[layer["op"]])
                    for k in layer:
                        if k == "op":
                            continue
                        set_sub_attrs(origin_op, k, layer[k])
                    inst_layers.append(origin_op)
            else:
                raise NotImplementedError(f"Only support str or dict format for layers, but got {layer}!")
        return inst_layers

    else:
        return raw_layers

def get_SPS_modelJson(modelCaseOrigin:str) -> dict:
    modelCase = modelCaseOrigin.lower()
    with open("examples/SPS_LLM_models.json", 'r') as f:
        pre_defined_cases = json.load(f)
        logger.info(f"Pre-defined cases: {pre_defined_cases.keys()}")
        if modelCase in pre_defined_cases:
            modelJson = pre_defined_cases[modelCase]
        else:
            modelJsonPath = f"examples/{modelCaseOrigin}.json"
            if os.path.exists(modelJsonPath):
                with open(modelJsonPath, 'r') as f:
                    modelJson = json.load(f)
            else:
                raise NotImplementedError(f"Model case {modelCase} is not supported or model file {modelJsonPath} not exists!")
        
    if "layers" not in modelJson:
        with open("examples/SPS_LLM_layers.json", 'r') as f1:
            pre_defined_layers = json.load(f1)
            
        for k in pre_defined_layers["LLM_layer_map"]:
            if ( k.endswith('*') and k[:-1] in modelCase ) or ( k == modelCase ):
                modelJson["layers"] = pre_defined_layers[pre_defined_layers["LLM_layer_map"][k]].copy()
                logger.info(f"Loaded pre-defined model layers: {pre_defined_layers['LLM_layer_map'][k]}")
                break

        if "layers" not in modelJson:
            raise NotImplementedError(f"Can't find pre-defined model layers for {modelCase}!")
        elif not isinstance(modelJson["layers"], list):
            raise NotImplementedError(f"Only support list format for layers in {modelCase}!")

        modelJson["layers"] = get_SPS_LLM_op(modelJson["layers"])
    return modelJson

class SPS_Model(superParSelBigOpBase):
    """
        LLM model for super parameter selection.
        Args:
            tpuJson   (str):   TPU json file.
            modelJson (str): Model json file.
    """
    type      :str  = "LLM" # Model type.
    moreAttr  :bool = True # More attribute not pre-defined.
    tpuJson   :str  = "" # TPU json file.
    modelJson :str  = "" # Model json file.
    
    seq_len   :int  = 1 # Sequence length.
    token_size:int  = 1 # Token size.
    layers    :list[dict] = [] # Model origin layer info.

    hidden_size: Optional[int]
    # t_post_process: float = 0 # Post process time.

    def __init__(self, **kwargs) -> None:        
        all_info = {}
        for s in ("modelJson", "tpuJson"):
            if s not in kwargs:
                raise NotImplementedError(f"{s} is not provided! {kwargs=}")
            if isinstance(kwargs[s], str): # Should be model json file path
                if not os.path.exists(kwargs[s]):
                    raise NotImplementedError(f"{s} {kwargs[s]} not exists!")
                with open(kwargs[s], 'r') as f:
                    all_info.update(json.load(f))
            elif isinstance(kwargs[s], dict): # Json info
                all_info.update(kwargs[s])
            else:
                raise NotImplementedError(f"{s} should be str or dict! {kwargs[s]=}")
        all_info.update(kwargs)

        if "layers" not in all_info:
            raise NotImplementedError("No layers specified!")
        
        if "hidden_size" not in all_info and "head" in all_info and "block_size" in all_info:
            self.hidden_size = all_info['head'] * all_info['block_size']
            logger.info(f"Setting {self.hidden_size=} : {all_info['head']} * {all_info['block_size']}")

        if "block_size" not in all_info and "hidden_size" in all_info and "head" in all_info:
            self.block_size = all_info['hidden_size'] // all_info['head']
            logger.info(f"Setting {self.block_size=} : {all_info['hidden_size']} / {all_info['head']}")

        ### TPU attributes should be set before SPS_OPs
        TPUAttrs.update(all_info)

        super().__init__(all_info)
        logger.info("Model loaded!")

        self.tp    = 'self.tp'
        self.batch = 'self.batch'
        self.seq_len = 1 if self.mode == 'decode' else self.token_size
        logger.info(f"Setting {self.mode=}, {self.seq_len=}")
        self.fullInstantiate(self.layers)

        for op_info in self.layers:
            if self.debug>0:
                logger.debug(f"instantiate {op_info}")
                if self.debug>1:
                    op_info['debug'] = self.debug
            self.sub_ops.append(gen_SPS_OP(op_info))
        logger.info("Model ops generated!\n")

    def backward(self):
        res = zero_res.copy()

        # CE loss, suppose that CE is fused op (only access data once) [from Chunlei.Zhang]
        lm_head_matrix_size = self.getMatrixSize(self.sub_ops[-1].out)
        tmp_io = 2 * 4 * lm_head_matrix_size / MB
        res["IOsize"] += tmp_io
        res['t_DMA']  += tmp_io / KB / TPUAttrs.DRAMBW * 1e6
        logger.debug(f"{self.name}.t_DMA += 2 * 4 * {lm_head_matrix_size} / KB / {TPUAttrs.DRAMBW} * 1e6 = {res['t_DMA']}")

        # Update all trainable parameters
        tmp_io = 2 * self.getNparam() / MB
        res["IOsize"] += tmp_io
        res['t_DMA']  += tmp_io / KB / TPUAttrs.DRAMBW * 1e6
        logger.debug(f"{self.name}.t_DMA += 2 * {self.getNparam()} / KB / {TPUAttrs.DRAMBW} * 1e6 = {res['t_DMA']}")

        return res

    def instantiateStr(self, s:str):
        """
            Instantiate string.
        """
        if isinstance(s, str):
            import re
            new_s = re.sub(r'\b([a-zA-Z_]\w*)\b', lambda m: str(getattr(self, m.group(1), m.group(1))), s)
            try:
                return eval(new_s)
            except:
                return new_s
            # return int(new_s) if new_s.isdigit() else new_s
        else:
            return s

    def fullInstantiate(self, info:dict|list):
        """
            Instantiate all attributes.
        """
        if isinstance(info, dict):
            for k, v in info.items():
                if isinstance(v, str):
                    info[k] = self.instantiateStr(v)
                elif isinstance(v, dict|list):
                    self.fullInstantiate(v)

        elif isinstance(info, list):
            for i, v in enumerate(info):
                if isinstance(v, str):
                    info[i] = self.instantiateStr(v)
                elif isinstance(v, dict|list):
                    self.fullInstantiate(v)

    def to_dict(self):
        res = {
            "name": self.name,
            "mode": self.mode,
            "tp": self.tp,
            "batch": self.batch,
            "seq_len": self.seq_len,
            "token_size": self.token_size,
            "DRAMBW": TPUAttrs.DRAMBW,
            "Nparam": self.getNparam(),
            "sub_ops": [],
        }
        for op in self.sub_ops:
            res["sub_ops"].append({"name": op.name, **op.to_dict()})
        res['model'] = self.last_res
        res['FTL'] = self.FTL
        res['TPS'] = self.TPS
        res['total_mem'] = self.total_mem
        return res

    @property
    def FTL(self):
        return self.last_res['t_MAX'] / 1e3
    
    @property
    def TPS(self):
        return self.batch*self.seq_len*1e6/self.last_res['t_MAX']
    
    @property
    def total_mem(self):
        return (self.last_res['m_act'] + self.last_res['m_weight'] + self.last_res['m_cache'] + self.last_res['m_other'])/KB

    def dump2excel(self, workbook:openpyxl.Workbook=None, outname:str="") -> None:  
        if isinstance(workbook, openpyxl.Workbook):
            for op in self.sub_ops:
                op.dump2excel(workbook)

            ### TPS and memory
            ws = workbook.active
            ws.append(["",]*16 + ["FTL [ms]", str(self.FTL), "TPS", str(self.TPS), "总内存 [GB]", str(self.total_mem)])
        super().dump2excel(workbook, outname if outname else f"SPS_{self.name}.xlsx")

def main(args=None):

    import argparse
    parser = argparse.ArgumentParser(description='Super parameter selection tool. Support override parameters in JSON. (--DRAMBW 491.4)')
    parser.add_argument('--case',  type=str, default='MM2',    help='Name of super parameter selection operator.')
    parser.add_argument('--mode',  type=str, default='decode', help='Mode of super parameter selection.')
    parser.add_argument('--tp',    type=int, default=1,        help='TP.')
    parser.add_argument('--batch', type=int, default=1,        help='Batch size.')
    parser.add_argument('--debug', type=int, default=0,        help='Debug mode.')
    parser.add_argument('--quant_method', '-q', type=str, default='f16', help='Quant method.')
    args, unknown_args = parser.parse_known_args(args=args)

    if args.debug > 0:
        logger.root.setLevel(logger.DEBUG)
    elif args.debug < 0:
        logger.root.setLevel(logger.ERROR)

    override_info = {}
    for i in range(0, len(unknown_args), 2):
        k = unknown_args[i].lstrip('-')
        if k=='s':
            k = 'token_size'
        v = unknown_args[i+1]
        try:
            v = int(v)
        except ValueError:
            try:
                v = float(v)
            except ValueError:
                logger.warning(f"{v} is not a number!")
        logger.info(f"setting: {k} = {v}")
        override_info[k] = v

    if args.case == "MM2":
        # KV_b = {
        #     "OM": False,
        #     "type": "MM2",
        #     "isParam": False,
        #     "matrix":{
        #         "in0": ["self.batch", 4096, 128,  64],
        #         "in1": [           1,  128,  64, 512],
        #         "out": ["self.batch", 4096, 128, 512]
        #     }
        # }
        # op = gen_SPS_OP(KV_b)
        # KV_b = {
        #     "type": "MM2",
        #     "matrix":{
        #         "in0": ["self.batch", 8192, 1, 8192],
        #         "in1": [           1, 1, 8192, 8192],
        #         "out": ["self.batch", 8192, 1, 8192]
        #     }
        # }
        KV_b = {
            "type": "MM2",
            "matrix":{
                "in0": ["self.batch", 1, 1, 3584],
                "in1": [           1, 1, 3584, 4608],
                "out": ["self.batch", 1, 1, 4608]
            }
        }
        op = gen_SPS_OP(KV_b)
        logger.info(op(args.mode, args.tp, args.batch))

    elif args.case == "Mix":
        KV_a_norm = {
            "OM": False,
            "type": "Mix",
            "matrix":{
                "in0": ["self.batch", 1, 4096, 512],
                "in1": [           1, 1,    1, 512],
                "out": ["self.batch", 1, 4096, 512]
            }
        }
        op = gen_SPS_OP(KV_a_norm)
        logger.info(op(args.mode, args.tp, args.batch))

    elif args.case == "Add":
        Add = {
            "type": "AR",
            "matrix": ["self.batch", 1, 4096, 8192]
        }
        op = gen_SPS_OP(Add)
        logger.info(op(args.mode, args.tp, args.batch))

    elif args.case == "C2C":
        C2C = {
            "type": "AllReduce",
            "matrix": ["self.batch", 1, 1, 16384]
            # "matrix": ["self.batch", 1, 4096, 7168]
        }
        op = gen_SPS_OP(C2C)
        logger.info(op(args.mode, args.tp, args.batch))

    elif args.case == "ATT":
        ATT_QKV = {
            "type": "Big",
            "cache":{
                "K_cache": ["self.batch", "128/self.tp", 4096, 128],
                "V_cache": ["self.batch", "128/self.tp", 4096, 128]
            },
                
            "matrix":{
                "in0": ["self.batch", 1, 4096, "128*(128+2*8)/self.tp"],
                "in1": [],
                "out": ["self.batch", 1, 4096, 8192]
            },
            
            "sub_info":[
                {
                    "name": "Neg_Q",
                    "NM": True,
                    "type":"AR",
                    "matrix": [ "self.batch", "32/self.tp", 1, "128/2" ]
                },
            ]
        }
        op = gen_SPS_OP(ATT_QKV)
        logger.info(op(args.mode, args.tp, args.batch))

    elif args.case == "Conv3D":
        Conv3D = {
            "type": "ConvND",
            "matrix":{
                "in0": ["2116*self.batch",    3, 2, 14, 14],
                "in1": [             1280,    3, 2, 14, 14],
                "out": ["2116*self.batch", 1280, 1,  1,  1]
            }
        }
        op = gen_SPS_OP(Conv3D)
        logger.info(op(args.mode, args.tp, args.batch))

    else:
        os.chdir(os.path.dirname(os.path.abspath(__file__)))
        tpuJson = "examples/tpu.json"
        modelJson = get_SPS_modelJson(args.case)

        op = SPS_Model(**args.__dict__, **override_info, tpuJson=tpuJson, modelJson=modelJson)
        run_cmd = {"quant_method": args.quant_method}
        logger.info(f"Running {args.case} with {args.mode} mode, {args.tp} TP, {args.batch} batch, {args.quant_method} quant method, {args.debug} debug mode")
        logger.info(f"Result = {op(args.mode, args.tp, args.batch, run_cmd)}\n")
        logger.info(f"Trainable parameters per chip = {op.getNparam()/G:.2f} G")
        logger.info(f"Total memory = {op.total_mem:.2f} GB")
        logger.info(f"Latency = {op.FTL:.2f} ms")
        logger.info(f"TPS = {op.TPS:.2f}")

    if args.debug == 2:
        op.dump2excel(None, f"SPS_{args.case}_tp{args.tp}_batch{args.batch}_seq{override_info.get('token_size', '')}.xlsx")
    elif args.debug == 3:
        with open(f"SPS_{op.name}_seq{override_info['token_size']}_{args.tp}TP_{args.batch}BS_{args.quant_method}_BW{override_info.get('DRAMBW', TPUAttrs.DRAMBW)}_{args.mode}.json", 'w') as f:
            f.write(json.dumps(op.to_dict(), indent=4))

    return op.to_dict()

if __name__ == "__main__":
    main()

# python superParTools.py --case llama3-70b --token_size 4096 --mode prefill --tp 8 --batch 16 --debug 2
